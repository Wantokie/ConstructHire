import io
import os
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'jobapp_secret_key')

# --- CONFIGURATION ---
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = 'static/uploads'

MAIL_SENDER = os.environ.get('MAIL_USERNAME', 'razonprinceeinstein@gmail.com')
SENDGRID_API_KEY = os.environ.get('SENDGRID_API_KEY')

db = SQLAlchemy(app)

if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

# --- SENDGRID HELPER ---
def send_email(to_email, subject, body):
    try:
        message = Mail(
            from_email=MAIL_SENDER,
            to_emails=to_email,
            subject=subject,
            plain_text_content=body
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        sg.send(message)
        print(f"Email sent to {to_email}")
    except Exception as e:
        print(f"SendGrid error: {e}")

# --- MODELS ---
class Job(db.Model):
    __tablename__ = 'jobs'
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(255), nullable=False)
    location = db.Column(db.String(255))
    pay = db.Column(db.String(100))
    description = db.Column(db.Text)
    job_type = db.Column(db.String(50))
    company = db.Column(db.String(100), default='ConstructHire')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    applications = db.relationship('Application', backref='job', cascade="all, delete-orphan")

class Application(db.Model):
    __tablename__ = 'applications'
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey('jobs.id'), nullable=False)
    applicant_name = db.Column(db.String(100), nullable=False)
    applicant_email = db.Column(db.String(100), nullable=False)
    contact_number = db.Column(db.String(20))
    address = db.Column(db.Text)
    experience_years = db.Column(db.String(50))
    resume_path = db.Column(db.String(255))
    status = db.Column(db.String(50), default='Pending')
    hired_at = db.Column(db.DateTime, nullable=True)

# --- ROUTES ---
@app.route('/')
def index():
    return render_template('home.html')

@app.route('/index')
def login_page():
    return render_template('index.html')

@app.route('/login', methods=['POST'])
def login():
    from sqlalchemy import text
    email = request.form.get('email')
    password = request.form.get('password')
    result = db.session.execute(
        text("SELECT * FROM admins WHERE email = :email AND password = :password"),
        {"email": email, "password": password}
    ).fetchone()
    if result:
        session['admin_logged_in'] = True
        session['admin_email'] = email
        return redirect(url_for('dashboard_view'))
    else:
        return redirect('/index?error=Invalid+email+or+password')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/admin')
@app.route('/dashboard')
def dashboard_view():
    return render_template('dashboard.html')

@app.route('/api/jobs', methods=['GET', 'POST'])
def handle_jobs():
    if request.method == 'POST':
        data = request.get_json()
        new_job = Job(
            title=data.get('title'),
            location=data.get('location'),
            pay=data.get('pay'),
            description=data.get('description'),
            job_type=data.get('type', 'Full-time')
        )
        db.session.add(new_job)
        db.session.commit()
        return jsonify({"message": "Job Created"}), 201

    jobs = Job.query.order_by(Job.created_at.desc()).all()
    return jsonify([{
        "id": j.id, "title": j.title, "location": j.location, "pay": j.pay,
        "company": j.company, "job_type": j.job_type, "description": j.description
    } for j in jobs])

@app.route('/api/applications', methods=['GET', 'POST'])
def handle_apps():
    if request.method == 'POST':
        f = request.form
        import re

        # --- Email: only @gmail.com and @yahoo.com ---
        email = f.get('email', '')
        if not re.match(r'^[^\s@]+@(gmail|yahoo)\.com$', email, re.IGNORECASE):
            return jsonify({"error": "Only @gmail.com or @yahoo.com emails are accepted."}), 400

        # --- Phone: numbers only ---
        phone = f.get("phone", "")
        if not re.match(r"^[0-9]+$", phone):
            return jsonify({"error": "Contact number must contain numbers only."}), 400

        # --- Block applications for Filled jobs ---
        job_id = f.get('job_id')
        job = Job.query.get(job_id)
        if job and job.job_type == 'Filled':
            return jsonify({"error": "This position is no longer accepting applications."}), 400

        resume = request.files.get('resume')
        filename = secure_filename(resume.filename) if resume else None
        if resume:
            resume.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        new_app = Application(
            job_id=f.get('job_id'), applicant_name=f.get('name'),
            applicant_email=f.get('email'), contact_number=f.get('phone'),
            address=f.get('address'), experience_years=f.get('experience'),
            resume_path=filename
        )
        db.session.add(new_app)
        db.session.commit()
        return jsonify({"message": "Sent"}), 201

    results = db.session.query(Application, Job.title).join(Job, Application.job_id == Job.id).all()
    return jsonify([{
        "id": a.id, "applicant_name": a.applicant_name, "applicant_email": a.applicant_email,
        "contact": a.contact_number, "address": a.address,
        "job_title": j_title, "status": a.status or 'Pending',
        "resume": a.resume_path, "experience": a.experience_years,
        "hired_at": a.hired_at.strftime('%Y-%m-%d %H:%M') if a.hired_at else None
    } for a, j_title in results])

@app.route('/api/applications/approve/<int:id>', methods=['POST'])
def approve_app(id):
    result = db.session.query(Application, Job.title).join(Job, Application.job_id == Job.id).filter(Application.id == id).first_or_404()
    app_obj, job_title = result
    app_obj.status = 'Completed'
    db.session.commit()

    send_email(
        to_email=app_obj.applicant_email,
        subject="Application Is On Process — ConstructHire",
        body=(
            f"Hi {app_obj.applicant_name},\n\n"
            f"Great news! Your application for the {job_title} position has been approved.\n\n"
            f"Prepare on your interview please bring the required documents including the Resume, "
            f"Barangay & Police Clearance and Valid Id\n\n"
            f"Come to our office at 9:00 am Located at City Of San Fernando Pampanga ConstructHire Main Office.\n"
            f"Good Luck on the interview!\n\n"
            f"Best regards,\nConstructHire HR department"
        )
    )

    return jsonify({"success": True})

@app.route('/api/applications/hire/<int:id>', methods=['POST'])
def hire_app(id):
    """Officially hire an approved (Completed) applicant and notify them by email."""
    result = db.session.query(Application, Job.title, Job.location, Job.pay).join(
        Job, Application.job_id == Job.id
    ).filter(Application.id == id).first_or_404()

    app_obj, job_title, job_location, job_pay = result

    if app_obj.status not in ('Completed', 'Hired'):
        return jsonify({"success": False, "error": "Applicant must be approved before hiring."}), 400

    app_obj.status = 'Hired'
    app_obj.hired_at = datetime.utcnow()
    db.session.commit()

    send_email(
        to_email=app_obj.applicant_email,
        subject="Congratulations! You're Hired — ConstructHire",
        body=(
            f"Dear {app_obj.applicant_name},\n\n"
            f"We are thrilled to officially offer you the position of {job_title} at ConstructHire!\n\n"
            f"Position Details:\n"
            f"  Job Title : {job_title}\n"
            f"  Location  : {job_location}\n"
            f"  Pay       : {job_pay}\n\n"
            f"Please expect a follow-up communication from our HR team with onboarding details, "
            f"your start date, and any documents you may need to prepare.\n\n"
            f"Welcome to the ConstructHire!\n\n"
            f"Warm regards,\nConstructHire HR department"
        )
    )

    return jsonify({"success": True})

@app.route('/api/applications/reject/<int:id>', methods=['POST'])
def reject_app(id):
    app_obj = Application.query.get_or_404(id)
    app_obj.status = 'Rejected'
    db.session.commit()
    return jsonify({"success": True})

# --- EXCEL EXPORT ---
@app.route('/api/applications/export', methods=['GET'])
def export_approved():
    """Export all Completed + Hired applicants to a formatted Excel file."""
    results = db.session.query(Application, Job.title, Job.location, Job.pay).join(
        Job, Application.job_id == Job.id
    ).filter(Application.status.in_(['Completed', 'Hired'])).order_by(Application.status, Application.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Applicants"

    # Styles
    header_font   = Font(name='Arial', bold=True, color='FFFFFF', size=11)
    header_fill   = PatternFill('solid', start_color='001736', end_color='001736')
    hired_fill    = PatternFill('solid', start_color='D4EDDA', end_color='D4EDDA')
    approved_fill = PatternFill('solid', start_color='FFF3CD', end_color='FFF3CD')
    center        = Alignment(horizontal='center', vertical='center')
    left          = Alignment(horizontal='left', vertical='center')
    thin          = Side(style='thin', color='CCCCCC')
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = 'ConstructHire — Approved & Hired Applicants'
    title_cell.font = Font(name='Arial', bold=True, size=14, color='001736')
    title_cell.alignment = center
    ws.row_dimensions[1].height = 30

    # Export date
    ws.merge_cells('A2:I2')
    ws['A2'].value = f'Exported on: {datetime.utcnow().strftime("%B %d, %Y %H:%M UTC")}'
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='64748B')
    ws['A2'].alignment = center
    ws.row_dimensions[2].height = 18

    # Blank row
    ws.row_dimensions[3].height = 8

    # Headers
    headers = ['#', 'Applicant Name', 'Email', 'Contact', 'Job Position',
               'Location', 'Pay', 'Experience', 'Status', 'Hired At']
    ws.append(headers)
    header_row = ws.max_row
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[header_row].height = 22

    # Data rows
    for i, (a, j_title, j_location, j_pay) in enumerate(results, 1):
        row_fill = hired_fill if a.status == 'Hired' else approved_fill
        row_data = [
            i,
            a.applicant_name,
            a.applicant_email,
            a.contact_number or '—',
            j_title,
            j_location or '—',
            j_pay or '—',
            a.experience_years or '—',
            a.status,
            a.hired_at.strftime('%Y-%m-%d %H:%M') if a.hired_at else '—'
        ]
        ws.append(row_data)
        r = ws.max_row
        ws.row_dimensions[r].height = 18
        for col in range(1, len(row_data) + 1):
            cell = ws.cell(row=r, column=col)
            cell.fill = row_fill
            cell.border = border
            cell.font = Font(name='Arial', size=10)
            cell.alignment = center if col == 1 else left

    # Column widths
    col_widths = [5, 22, 28, 16, 20, 20, 18, 14, 12, 18]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Summary row
    ws.append([])
    ws.append(['', f'Total Records: {len(results)}',
               f'Hired: {sum(1 for a,*_ in results if a.status == "Hired")}',
               f'Approved: {sum(1 for a,*_ in results if a.status == "Completed")}'])
    summary_row = ws.max_row
    for col in range(2, 5):
        cell = ws.cell(row=summary_row, column=col)
        cell.font = Font(name='Arial', bold=True, size=10, color='001736')

    # Freeze header rows
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"approved_applicants_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/jobs/<int:id>', methods=['PUT'])
def update_job(id):
    job = Job.query.get_or_404(id)
    data = request.get_json()
    job.title = data.get('title', job.title)
    job.location = data.get('location', job.location)
    job.pay = data.get('pay', job.pay)
    job.description = data.get('description', job.description)
    job.job_type = data.get('type', job.job_type)
    db.session.commit()
    return jsonify({"message": "Job Updated"})

@app.route('/api/jobs/<int:id>', methods=['DELETE'])
def delete_job(id):
    job = Job.query.get_or_404(id)
    db.session.delete(job)
    db.session.commit()
    return jsonify({"message": "Job Deleted"})

@app.route('/api/jobs/<int:id>/filled', methods=['POST'])
def mark_job_filled(id):
    job = Job.query.get_or_404(id)
    job.job_type = 'Filled'
    db.session.commit()
    return jsonify({"message": "Job Marked as Filled"})

# --- AUTO CREATE TABLES ---
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True)